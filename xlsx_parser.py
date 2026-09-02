# -*- coding: utf-8 -*-
"""
XLSX 시트별 표를 grid(list of list)로 변환. merged_cells.ranges로 병합 구간을 확인해
병합의 원본 셀(좌상단) 값만 채우고 나머지는 merge_map=True로 표시한다.
"""
import openpyxl
from ..utils import clean_whitespace


def parse_xlsx_tables(file_obj) -> list:
    """
    반환: [
      {"sheet": str, "grid": [[cell_text,...],...], "merge_map": [[bool,...],...]}, ...
    ]
    시트 전체를 하나의 표로 취급한다(비교표가 보통 시트 전체를 차지하는 형태라고 가정).
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    results = []

    for sheet in wb.worksheets:
        max_row = sheet.max_row
        max_col = sheet.max_column
        if max_row == 0 or max_col == 0:
            continue

        grid = [["" for _ in range(max_col)] for _ in range(max_row)]
        merge_map = [[False for _ in range(max_col)] for _ in range(max_row)]

        merged_ranges = list(sheet.merged_cells.ranges)
        merged_absorbed = set()  # (row, col) 0-indexed, 병합에 흡수되어 origin이 아닌 셀
        for rng in merged_ranges:
            min_r, min_c = rng.min_row, rng.min_col
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if (r, c) != (min_r, min_c):
                        merged_absorbed.add((r, c))

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                if (r, c) in merged_absorbed:
                    merge_map[r - 1][c - 1] = True
                    continue
                cell = sheet.cell(row=r, column=c)
                grid[r - 1][c - 1] = clean_whitespace(cell.value if cell.value is not None else "")

        # 완전히 빈 시트는 제외
        if any(any(row) for row in grid):
            results.append({"sheet": sheet.title, "grid": grid, "merge_map": merge_map})

    return results
